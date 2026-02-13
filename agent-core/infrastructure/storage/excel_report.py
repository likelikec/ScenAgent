import os
import argparse
import json
import re
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config.settings import resolve_summary_llm_params

def _normalize_output_lang(output_lang):
    v = (output_lang or "").strip().lower()
    if v in ("zh", "ch", "cn", "zh-cn", "zh_hans", "zh-hans"):
        return "zh"
    return "en"

def _localize(output_lang, zh, en):
    return zh if output_lang == "zh" else en

def _headers_simple(output_lang):
    if output_lang == "zh":
        return ["测试用例编号", "模块名称", "用例说明", "执行步骤", "成功步骤", "输入数据", "执行结果", "执行原因", "token总数", "探索步骤数", "执行时间"]
    return ["Test Case ID", "Module Name", "Case Description", "Execution Steps", "Successful Steps", "Input Data", "Execution Result", "Execution Reason", "Total Tokens", "Exploration Steps", "Execution Time"]

def _headers_full(output_lang):
    return _headers_simple(output_lang)

def _ensure_wb_and_sheet(path, headers):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row >= 1:
            existing_headers = [c.value for c in ws[1]]
            desired = headers
            existing_str = [str(x).strip() if x is not None else "" for x in existing_headers]
            desired_str = [str(x).strip() if x is not None else "" for x in desired]
            
            mismatch = False
            if len(existing_str) < len(desired_str):
                mismatch = True
            else:
                if existing_str[:len(desired_str)] != desired_str:
                    mismatch = True
            
            if mismatch:
                for i, h in enumerate(desired, 1):
                    ws.cell(row=1, column=i, value=h)
                
                if ws.max_column > len(desired):
                    for col in range(len(desired) + 1, ws.max_column + 1):
                        ws.cell(row=1, column=col, value=None)
                
                wb.save(path)
        else:
            ws.append(headers)
            wb.save(path)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    wb.save(path)
    return wb, ws

def _sanitize_name(s):
    invalid = '<>:"/\\|?*'
    r = ''.join(ch for ch in (s or "") if ch not in invalid)
    r = r.strip()
    if len(r) == 0:
        r = "Unknown"
    return r

def _count_exploration_steps(steps_dir: str) -> int:
    if not steps_dir or not os.path.exists(steps_dir):
        return 0
    step_folders = []
    for d in os.listdir(steps_dir):
        if d.startswith("step_"):
            try:
                n = int(d.split("_")[-1])
            except Exception:
                n = 0
            step_folders.append((n, d))
    step_folders.sort(key=lambda x: x[0])
    if not step_folders:
        return 0
    total = len(step_folders)
    _, lastd = step_folders[-1]
    last_step_dir = os.path.join(steps_dir, lastd)
    has_task_judge = any(
        os.path.exists(os.path.join(last_step_dir, name))
        for name in ("task_judge.json", "task_judge.zh.json")
    )
    if has_task_judge:
        total -= 1
    return max(total, 0)

def _extract_thought(resp):
    if "### Thought" in resp and "### Plan" in resp:
        return resp.split("### Thought")[-1].split("### Plan")[0].strip()
    return ""

def _extract_error(resp):
    if "### Error Description" in resp:
        return resp.split("### Error Description")[-1].strip()
    return ""

def _parse_dt(s):
    t = (s or "").strip()
    if not t:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(t, fmt)
        except Exception:
            pass
    return None

def _number_steps(s):
    t = (s or "").strip()
    if not t:
        return t
    t = t.replace("Finished", "").strip()
    parts = []
    tmp = t.replace("\r\n", "\n").replace("\r", "\n")
    numbered_pattern = r'\d+\.\s+(.+?)(?=\s+\d+\.\s+|$)'
    matches = re.findall(numbered_pattern, tmp, re.DOTALL)
    if matches:
        for match in matches:
            content = match.strip()
            if content:
                parts.append(content)
    elif tmp.startswith("- ") or " - " in tmp:
        tmp = tmp.replace("\n- ", " - ").replace("\n - ", " - ")
        for p in tmp.split(" - "):
            ps = p.strip()
            if ps:
                ps = re.sub(r'^\d+\.\s*', '', ps).strip()
                if ps:
                    parts.append(ps)
    else:
        for ln in re.split(r"[\r\n]+", tmp):
            ps = ln.strip(" -")
            if ps:
                ps = re.sub(r'^\d+\.\s*', '', ps).strip()
                if ps:
                    parts.append(ps)
    return "\n".join(f"{i+1}. {parts[i]}" for i in range(len(parts)))

def _read_json_file_safely(file_path, max_retries=3, retry_delay=0.5):
    for attempt in range(max_retries):
        if not os.path.exists(file_path):
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                continue
            return None
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data
        except (json.JSONDecodeError, IOError, ValueError):
            if attempt == max_retries - 1:
                return None
            time.sleep(retry_delay)
    return None

def _get_total_plan_from_files(run_dir):
    exec_steps = ""
    script_path = os.path.join(run_dir, "script.json")
    if os.path.exists(script_path):
        sj = _read_json_file_safely(script_path)
        if sj:
            exec_steps = sj.get("total_plan") or sj.get("goal") or ""
    if not exec_steps:
        infopool_path = os.path.join(run_dir, "infopool.json")
        if os.path.exists(infopool_path):
            ip = _read_json_file_safely(infopool_path)
            if ip:
                exec_steps = ip.get("total_plan") or ""
    return exec_steps

def _filter_successful_steps(
    exec_steps,
    output_lang,
    vllm_api_key=None,
    vllm_base_url=None,
    vllm_model_name=None,
    summary_api_key=None,
    summary_base_url=None,
    summary_model_name=None,
):
    """使用LLM筛选成功步骤"""
    if not exec_steps:
        return ""
        
    try:
        from infrastructure.llm.llm_factory import LLMFactory
        params = resolve_summary_llm_params(
            vllm_api_key=vllm_api_key,
            vllm_base_url=vllm_base_url,
            vllm_model_name=vllm_model_name,
            summary_api_key=summary_api_key,
            summary_base_url=summary_base_url,
            summary_model_name=summary_model_name,
        )
        provider_type = params.get("provider_type") or "gui_owl"
        api_key = params.get("api_key")
        base_url = params.get("base_url")
        model_name = params.get("model_name")
        temperature = params.get("temperature", 0.0)
        max_retry = params.get("max_retry", 10)
        if not api_key:
            return exec_steps
        if not base_url or not model_name:
            return exec_steps
            
        llm = LLMFactory.create(
            provider_type=provider_type,
            api_key=api_key,
            base_url=base_url,
            model_name=model_name,
            temperature=temperature,
            max_retry=max_retry,
        )
        
        if output_lang == "zh":
            prompt = f"""你是一个专业的测试步骤分析助手。请分析以下执行步骤历史，筛选出成功到达最终目的地的有效步骤序列。

### 原始执行步骤 ###
{exec_steps}

### 筛选要求 ###
1. 去除所有尝试性、探索性但最终未成功或被回退的步骤。
2. 去除所有重复或冗余的操作。
3. 仅保留直接贡献于最终目标达成的关键步骤。
4. 保持步骤的原始顺序。
5. **严禁无中生有**：输出的步骤必须在原始步骤中存在，不得编造。
6. 输出格式必须为编号列表：
   1. 第一步
   2. 第二步
   ...

### 输出 ###
请直接输出筛选后的步骤列表，不要包含任何其他解释或前缀后缀。
"""
        else:
            prompt = f"""You are a professional test-step analysis assistant. Analyze the execution steps and keep only the effective steps that contributed to reaching the final goal successfully.

### Raw Execution Steps ###
{exec_steps}

### Requirements ###
1. Remove exploratory/failed/backtracked steps that did not lead to success.
2. Remove duplicates and redundant actions.
3. Keep only key steps that directly contribute to achieving the final goal.
4. Keep the original order.
5. Do NOT fabricate: every output step must come from the raw steps.
6. Output must be a numbered list:
   1. First step
   2. Second step
   ...

### 输出 ###
Output only the filtered numbered steps, without any extra text.
"""
        response, _, _ = llm.predict_mm(prompt, [])
        return _number_steps(response)
    except Exception as e:
        print(f"Failed to filter steps with LLM: {e}")
        return exec_steps

def generate_excel_report(
    scenario_file,
    app_id,
    scenario_id,
    logs_root="./output",
    output_excel_path="./reports/report.xlsx",
    output_lang="zh",
    vllm_api_key=None,
    vllm_base_url=None,
    vllm_model_name=None,
    summary_api_key=None,
    summary_base_url=None,
    summary_model_name=None,
):
    output_lang = _normalize_output_lang(output_lang)
    if not os.path.exists(scenario_file):
        raise ValueError("scenario_file not found")
    with open(scenario_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    apps = data.get("apps") or []
    scenarios = data.get("scenarios") or []
    app_name = None
    for a in apps:
        if a.get("id") == app_id:
            app_name = a.get("name")
            break
    if app_name is None and apps:
        app_name = apps[0].get("name")
    scd = None
    for s in scenarios:
        if s.get("id") == scenario_id:
            scd = s
            break
    if scd is None and scenarios:
        scd = scenarios[0]
    module_name = scd.get("name") if scd else ""
    case_desc = scd.get("description") if scd else ""
    extra_info_obj = scd.get("extra-info") if scd else {}
    extra_info_str = json.dumps(extra_info_obj, ensure_ascii=False)
    safe_app = _sanitize_name(app_name)
    safe_sce = _sanitize_name(module_name)
    if not os.path.exists(logs_root):
        raise ValueError("logs_root not found")
    candidates = []
    for d in os.listdir(logs_root):
        p = os.path.join(logs_root, d)
        if os.path.isdir(p) and d.startswith(f"T-{safe_app}-{safe_sce}-"):
            candidates.append((p, os.path.getmtime(p)))
    if not candidates:
        for d in os.listdir(logs_root):
            p = os.path.join(logs_root, d)
            if os.path.isdir(p) and d.startswith("T-"):
                candidates.append((p, os.path.getmtime(p)))
    if not candidates:
        raise ValueError("no run directories found")
    candidates.sort(key=lambda x: x[1], reverse=True)
    run_dir = candidates[0][0]
    case_id = os.path.basename(run_dir)
    steps_dir = os.path.join(run_dir, "Steps")
    step_folders = []
    if os.path.exists(steps_dir):
        for d in os.listdir(steps_dir):
            if d.startswith("step_"):
                try:
                    n = int(d.split("_")[-1])
                except Exception:
                    n = 0
                step_folders.append((n, d))
    step_folders.sort(key=lambda x: x[0])
    last_planner_thought = ""
    last_error_desc = ""
    if step_folders:
        n, lastd = step_folders[-1]
        mgr = os.path.join(steps_dir, lastd, "planner.zh.json" if output_lang == "zh" else "planner.json")
        ref = os.path.join(steps_dir, lastd, "reflector.zh.json" if output_lang == "zh" else "reflector.json")
        if os.path.exists(mgr):
            with open(mgr, "r", encoding="utf-8") as f:
                mj = json.load(f)
            last_planner_thought = _extract_thought(mj.get("response") or "")
        if os.path.exists(ref):
            with open(ref, "r", encoding="utf-8") as f:
                rj = json.load(f)
            last_error_desc = _extract_error(rj.get("response") or "")
    exploration_steps = _count_exploration_steps(steps_dir)
    task_results_path = os.path.join(run_dir, "task_results.json")
    step_limit = 1.0
    task_status = ""
    test_status_report = ""
    total_tokens = 0
    execution_steps = 0
    duration_seconds = ""
    if os.path.exists(task_results_path):
        with open(task_results_path, "r", encoding="utf-8") as f:
            tr = json.load(f)
        step_limit = tr.get("step_limit", tr.get("hit_step_limit", 1.0))
        task_status = tr.get("task_status", "")
        if output_lang == "zh":
            test_status_report = (
                tr.get("test_status_report_zh")
                or tr.get("test_status_report", "")
                or tr.get("status_reason_zh")
                or tr.get("status_reason", "")
            )
        else:
            test_status_report = tr.get("test_status_report", "") or tr.get("status_reason", "")
        total_tokens = int(tr.get("total_tokens") or 0)
        execution_steps = int(tr.get("execution_steps") or 0)
        start_dt = _parse_dt(tr.get("start_dtime") or "")
        finish_dt = _parse_dt(tr.get("finish_dtime") or "")
        if start_dt and finish_dt:
            duration_seconds = round((finish_dt - start_dt).total_seconds(), 3)
    exec_steps = _get_total_plan_from_files(run_dir)
    completed = False if (step_limit and float(step_limit) != 0.0) else True
    if isinstance(task_status, str) and task_status.strip().lower() in ("not completed", "not_completed"):
        completed = False
    elif isinstance(task_status, str) and task_status.strip().lower() in ("completed",):
        completed = True
    exec_result = _localize(output_lang, "测试完成", "Completed") if completed else _localize(output_lang, "测试未完成", "Not Completed")
    exec_reason = test_status_report or ""
    if not exec_reason and not completed:
        exec_reason = _localize(output_lang, "达到最大执行次数", "Reached maximum execution limit")
    
    # 处理执行步骤
    formatted_exec_steps = _number_steps(exec_steps)
    success_steps = (
        _filter_successful_steps(
            formatted_exec_steps,
            output_lang,
            vllm_api_key=vllm_api_key,
            vllm_base_url=vllm_base_url,
            vllm_model_name=vllm_model_name,
            summary_api_key=summary_api_key,
            summary_base_url=summary_base_url,
            summary_model_name=summary_model_name,
        )
        if completed
        else ""
    )
    
    headers = _headers_simple(output_lang)
    row = [case_id, module_name or "", case_desc or "", formatted_exec_steps or "", success_steps, extra_info_str or "", exec_result, exec_reason, total_tokens, exploration_steps, duration_seconds]
    out_dir = os.path.dirname(output_excel_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)
    wb, ws = _ensure_wb_and_sheet(output_excel_path, headers)
    ws.append(row)
    wb.save(output_excel_path)

def write_report_for_run(
    run_dir,
    scenario_file,
    app_id,
    scenario_id,
    output_lang="zh",
    vllm_api_key=None,
    vllm_base_url=None,
    vllm_model_name=None,
    summary_api_key=None,
    summary_base_url=None,
    summary_model_name=None,
):
    output_lang = _normalize_output_lang(output_lang)
    if not os.path.exists(run_dir):
        raise ValueError("run_dir not found")
    logs_root = os.path.dirname(run_dir)
    output_excel_path = os.path.join(logs_root, "results.xlsx")
    if not os.path.exists(scenario_file):
        raise ValueError("scenario_file not found")
    with open(scenario_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    apps = data.get("apps") or []
    scenarios = data.get("scenarios") or []
    app_name = None
    for a in apps:
        if a.get("id") == app_id:
            app_name = a.get("name")
            break
    if app_name is None and apps:
        app_name = apps[0].get("name")
    scd = None
    for s in scenarios:
        if s.get("id") == scenario_id:
            scd = s
            break
    if scd is None and scenarios:
        scd = scenarios[0]
    module_name = scd.get("name") if scd else ""
    case_desc = scd.get("description") if scd else ""
    extra_info_obj = scd.get("extra-info") if scd else {}
    extra_info_str = json.dumps(extra_info_obj, ensure_ascii=False)
    case_id = os.path.basename(run_dir)
    steps_dir = os.path.join(run_dir, "Steps")
    step_folders = []
    if os.path.exists(steps_dir):
        for d in os.listdir(steps_dir):
            if d.startswith("step_"):
                try:
                    n = int(d.split("_")[-1])
                except Exception:
                    n = 0
                step_folders.append((n, d))
    step_folders.sort(key=lambda x: x[0])
    last_planner_thought = ""
    last_error_desc = ""
    if step_folders:
        n, lastd = step_folders[-1]
        mgr = os.path.join(steps_dir, lastd, "planner.zh.json" if output_lang == "zh" else "planner.json")
        ref = os.path.join(steps_dir, lastd, "reflector.zh.json" if output_lang == "zh" else "reflector.json")
        if os.path.exists(mgr):
            with open(mgr, "r", encoding="utf-8") as f:
                mj = json.load(f)
            last_planner_thought = _extract_thought(mj.get("response") or "")
        if os.path.exists(ref):
            with open(ref, "r", encoding="utf-8") as f:
                rj = json.load(f)
            last_error_desc = _extract_error(rj.get("response") or "")
    exploration_steps = _count_exploration_steps(steps_dir)
    task_results_path = os.path.join(run_dir, "task_results.json")
    step_limit = 1.0
    task_status = ""
    test_status_report = ""
    total_tokens = 0
    execution_steps = 0
    duration_seconds = ""
    if os.path.exists(task_results_path):
        with open(task_results_path, "r", encoding="utf-8") as f:
            tr = json.load(f)
        step_limit = tr.get("step_limit", tr.get("hit_step_limit", 1.0))
        task_status = tr.get("task_status", "")
        if output_lang == "zh":
            test_status_report = (
                tr.get("test_status_report_zh")
                or tr.get("test_status_report", "")
                or tr.get("status_reason_zh")
                or tr.get("status_reason", "")
            )
        else:
            test_status_report = tr.get("test_status_report", "") or tr.get("status_reason", "")
        total_tokens = int(tr.get("total_tokens") or 0)
        execution_steps = int(tr.get("execution_steps") or 0)
        start_dt = _parse_dt(tr.get("start_dtime") or "")
        finish_dt = _parse_dt(tr.get("finish_dtime") or "")
        if start_dt and finish_dt:
            duration_seconds = round((finish_dt - start_dt).total_seconds(), 3)
    exec_steps = _get_total_plan_from_files(run_dir)
    completed = False if (step_limit and float(step_limit) != 0.0) else True
    if isinstance(task_status, str) and task_status.strip().lower() in ("not completed", "not_completed"):
        completed = False
    elif isinstance(task_status, str) and task_status.strip().lower() in ("completed",):
        completed = True
    exec_result = _localize(output_lang, "测试完成", "Completed") if completed else _localize(output_lang, "测试未完成", "Not Completed")
    exec_reason = test_status_report or ""
    if not exec_reason and not completed:
        exec_reason = _localize(output_lang, "达到最大执行次数", "Reached maximum execution limit")
    formatted_exec_steps = _number_steps(exec_steps)
    success_steps = (
        _filter_successful_steps(
            formatted_exec_steps,
            output_lang,
            vllm_api_key=vllm_api_key,
            vllm_base_url=vllm_base_url,
            vllm_model_name=vllm_model_name,
            summary_api_key=summary_api_key,
            summary_base_url=summary_base_url,
            summary_model_name=summary_model_name,
        )
        if completed
        else ""
    )
    headers = _headers_full(output_lang)
    row = [case_id, module_name or "", case_desc or "", formatted_exec_steps or "", success_steps, extra_info_str or "", exec_result, exec_reason, total_tokens, exploration_steps, duration_seconds]
    wb, ws = _ensure_wb_and_sheet(output_excel_path, headers)
    ws.append(row)
    wb.save(output_excel_path)

def _main():
    parser = argparse.ArgumentParser(description="Generate Excel report from Mobile-Agent-v4 run directories")
    parser.add_argument("--scenario_file", type=str, required=True)
    parser.add_argument("--app_id", type=str, required=True)
    parser.add_argument("--scenario_id", type=str, required=True)
    parser.add_argument("--logs_root", type=str, default="./output")
    parser.add_argument("--output_excel_path", type=str, default="./reports/report.xlsx")
    parser.add_argument("--output_lang", type=str, default="zh")
    parser.add_argument("--api_key", type=str)
    parser.add_argument("--base_url", type=str)
    parser.add_argument("--model", type=str)
    parser.add_argument("--summary_api_key", type=str)
    parser.add_argument("--summary_base_url", type=str)
    parser.add_argument("--summary_model", type=str)
    args = parser.parse_args()

    generate_excel_report(
        scenario_file=args.scenario_file,
        app_id=args.app_id,
        scenario_id=args.scenario_id,
        logs_root=args.logs_root,
        output_excel_path=args.output_excel_path,
        output_lang=args.output_lang,
        vllm_api_key=args.api_key,
        vllm_base_url=args.base_url,
        vllm_model_name=args.model,
        summary_api_key=args.summary_api_key,
        summary_base_url=args.summary_base_url,
        summary_model_name=args.summary_model,
    )

if __name__ == "__main__":
    _main()
